#!/usr/bin/env python3
"""
Portfolio Pro 数据导出模块
支持：Excel、CSV、JSON格式导出
"""
import json
import csv
import io
from datetime import datetime
from typing import Dict, List, Optional
from dataclasses import dataclass

# 可选依赖：openpyxl用于Excel导出
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("[Warning] openpyxl未安装，Excel导出将降级为CSV")

HOLDINGS_FILE = '/root/.openclaw/workspace/portfolio-tracker/data/holdings.json'
HISTORY_FILE = '/root/.openclaw/workspace/portfolio-tracker/data/history.json'

@dataclass
class ExportConfig:
    """导出配置"""
    format: str = 'excel'  # excel/csv/json
    include_stocks: bool = True
    include_funds: bool = True
    include_history: bool = True
    date_range: Optional[tuple] = None

class DataExporter:
    """数据导出器"""
    
    def __init__(self):
        self.holdings = self._load_json(HOLDINGS_FILE)
        self.history = self._load_json(HISTORY_FILE)
    
    def _load_json(self, filepath: str) -> Dict:
        """加载JSON文件"""
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            print(f"[Error] 加载失败 {filepath}: {e}")
            return {}
    
    def export_holdings_to_csv(self) -> str:
        """导出持仓为CSV"""
        output = io.StringIO()
        writer = csv.writer(output)
        
        # 写入表头
        writer.writerow(['代码', '名称', '类型', '持仓数量', '成本价', '当前价', 
                        '市值', '累计盈亏', '收益率%', '日涨跌%', '市场'])
        
        accounts = self.holdings.get('accounts', [])
        for account in accounts:
            account_name = account.get('name', '')
            
            # 股票
            for stock in account.get('stocks', []):
                writer.writerow([
                    stock.get('code', ''),
                    stock.get('name', ''),
                    '股票',
                    stock.get('shares', 0),
                    stock.get('cost', 0),
                    stock.get('price', 0),
                    stock.get('marketValue', 0),
                    stock.get('totalPnL', 0),
                    stock.get('returnRate', 0),
                    stock.get('dailyChange', 0),
                    stock.get('market', '')
                ])
            
            # 基金
            for fund in account.get('funds', []):
                writer.writerow([
                    fund.get('code', ''),
                    fund.get('name', ''),
                    '基金',
                    fund.get('shares', 0),
                    fund.get('cost', 0) / fund.get('shares', 1) if fund.get('shares', 0) > 0 else 0,
                    fund.get('nav', 0),
                    fund.get('marketValue', 0),
                    fund.get('totalPnL', 0),
                    fund.get('returnRate', 0),
                    fund.get('dailyChange', 0),
                    fund.get('market', '')
                ])
        
        return output.getvalue()
    
    def export_history_to_csv(self) -> str:
        """导出历史收益为CSV"""
        output = io.StringIO()
        writer = csv.writer(output)
        
        writer.writerow(['日期', '总资产', '累计收益', '今日盈亏', '沪深300对标', 
                        '风险等级', '最大回撤', '持仓集中度'])
        
        for record in self.history.get('records', []):
            writer.writerow([
                record.get('date', ''),
                record.get('totalAssets', 0),
                record.get('cumulativeReturn', 0),
                record.get('dailyPnL', 0),
                record.get('hs300Benchmark', 0),
                record.get('riskLevel', ''),
                record.get('maxDrawdown', 0),
                record.get('concentration', 0)
            ])
        
        return output.getvalue()
    
    def export_to_excel(self, output_path: str) -> bool:
        """导出为Excel文件"""
        if not EXCEL_AVAILABLE:
            # 降级为CSV
            return self._export_to_csv_files(output_path.replace('.xlsx', '.csv'))
        
        try:
            wb = Workbook()
            
            # 样式定义
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='6366F1', end_color='6366F1', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Sheet 1: 持仓明细
            ws1 = wb.active
            ws1.title = '持仓明细'
            
            headers = ['代码', '名称', '类型', '持仓数量', '成本价', '当前价', 
                      '市值', '累计盈亏', '收益率%', '日涨跌%', '市场']
            for col, header in enumerate(headers, 1):
                cell = ws1.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            row = 2
            accounts = self.holdings.get('accounts', [])
            for account in accounts:
                for stock in account.get('stocks', []):
                    ws1.cell(row=row, column=1, value=stock.get('code', ''))
                    ws1.cell(row=row, column=2, value=stock.get('name', ''))
                    ws1.cell(row=row, column=3, value='股票')
                    ws1.cell(row=row, column=4, value=stock.get('shares', 0))
                    ws1.cell(row=row, column=5, value=stock.get('cost', 0))
                    ws1.cell(row=row, column=6, value=stock.get('price', 0))
                    ws1.cell(row=row, column=7, value=stock.get('marketValue', 0))
                    ws1.cell(row=row, column=8, value=stock.get('totalPnL', 0))
                    ws1.cell(row=row, column=9, value=f"{stock.get('returnRate', 0):.2f}%")
                    ws1.cell(row=row, column=10, value=f"{stock.get('dailyChange', 0):.2f}%")
                    ws1.cell(row=row, column=11, value=stock.get('market', ''))
                    row += 1
                
                for fund in account.get('funds', []):
                    ws1.cell(row=row, column=1, value=fund.get('code', ''))
                    ws1.cell(row=row, column=2, value=fund.get('name', ''))
                    ws1.cell(row=row, column=3, value='基金')
                    ws1.cell(row=row, column=4, value=fund.get('shares', 0))
                    cost = fund.get('cost', 0) / fund.get('shares', 1) if fund.get('shares', 0) > 0 else 0
                    ws1.cell(row=row, column=5, value=cost)
                    ws1.cell(row=row, column=6, value=fund.get('nav', 0))
                    ws1.cell(row=row, column=7, value=fund.get('marketValue', 0))
                    ws1.cell(row=row, column=8, value=fund.get('totalPnL', 0))
                    ws1.cell(row=row, column=9, value=f"{fund.get('returnRate', 0):.2f}%")
                    ws1.cell(row=row, column=10, value=f"{fund.get('dailyChange', 0):.2f}%")
                    ws1.cell(row=row, column=11, value=fund.get('market', ''))
                    row += 1
            
            # 调整列宽
            for col in range(1, 12):
                ws1.column_dimensions[get_column_letter(col)].width = 15
            
            # Sheet 2: 历史收益
            ws2 = wb.create_sheet('历史收益')
            
            headers2 = ['日期', '总资产', '累计收益', '今日盈亏', '沪深300对标', 
                       '风险等级', '最大回撤', '持仓集中度']
            for col, header in enumerate(headers2, 1):
                cell = ws2.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            for row_idx, record in enumerate(self.history.get('records', []), 2):
                ws2.cell(row=row_idx, column=1, value=record.get('date', ''))
                ws2.cell(row=row_idx, column=2, value=record.get('totalAssets', 0))
                ws2.cell(row=row_idx, column=3, value=record.get('cumulativeReturn', 0))
                ws2.cell(row=row_idx, column=4, value=record.get('dailyPnL', 0))
                ws2.cell(row=row_idx, column=5, value=record.get('hs300Benchmark', 0))
                ws2.cell(row=row_idx, column=6, value=record.get('riskLevel', ''))
                ws2.cell(row=row_idx, column=7, value=record.get('maxDrawdown', 0))
                ws2.cell(row=row_idx, column=8, value=record.get('concentration', 0))
            
            # Sheet 3: 汇总信息
            ws3 = wb.create_sheet('资产汇总')
            
            # 计算汇总数据
            total_assets = 0
            total_pnl = 0
            daily_pnl = 0
            
            accounts = self.holdings.get('accounts', [])
            for account in accounts:
                for stock in account.get('stocks', []):
                    total_assets += stock.get('marketValue', 0)
                    total_pnl += stock.get('totalPnL', 0)
                    daily_pnl += stock.get('dailyPnL', 0)
                for fund in account.get('funds', []):
                    total_assets += fund.get('marketValue', 0)
                    total_pnl += fund.get('totalPnL', 0)
                    daily_pnl += fund.get('dailyPnL', 0)
            
            return_rate = (total_pnl / (total_assets - total_pnl) * 100) if (total_assets - total_pnl) > 0 else 0
            
            ws3.cell(row=1, column=1, value='指标').font = header_font
            ws3.cell(row=1, column=2, value='数值').font = header_font
            ws3.cell(row=1, column=1).fill = header_fill
            ws3.cell(row=1, column=2).fill = header_fill
            
            ws3.cell(row=2, column=1, value='总资产')
            ws3.cell(row=2, column=2, value=total_assets)
            
            ws3.cell(row=3, column=1, value='累计盈亏')
            ws3.cell(row=3, column=2, value=total_pnl)
            
            ws3.cell(row=4, column=1, value='今日盈亏')
            ws3.cell(row=4, column=2, value=daily_pnl)
            
            ws3.cell(row=5, column=1, value='收益率')
            ws3.cell(row=5, column=2, value=f"{return_rate:.2f}%")
            
            ws3.cell(row=6, column=1, value='数据更新时间')
            ws3.cell(row=6, column=2, value=self.holdings.get('lastUpdate', ''))
            
            # 保存
            wb.save(output_path)
            print(f"✅ Excel导出成功: {output_path}")
            return True
            
        except Exception as e:
            print(f"[Error] Excel导出失败: {e}")
            return self._export_to_csv_files(output_path.replace('.xlsx', ''))
    
    def _export_to_csv_files(self, base_path: str) -> bool:
        """降级为CSV导出"""
        try:
            # 持仓明细
            with open(f"{base_path}_holdings.csv", 'w', newline='', encoding='utf-8-sig') as f:
                f.write(self.export_holdings_to_csv())
            
            # 历史收益
            with open(f"{base_path}_history.csv", 'w', newline='', encoding='utf-8-sig') as f:
                f.write(self.export_history_to_csv())
            
            print(f"✅ CSV导出成功: {base_path}_holdings.csv, {base_path}_history.csv")
            return True
        except Exception as e:
            print(f"[Error] CSV导出失败: {e}")
            return False
    
    def export_to_json(self, output_path: str) -> bool:
        """导出为JSON文件"""
        try:
            export_data = {
                'exportTime': datetime.now().isoformat(),
                'holdings': self.holdings,
                'history': self.history
            }
            
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(export_data, f, ensure_ascii=False, indent=2)
            
            print(f"✅ JSON导出成功: {output_path}")
            return True
        except Exception as e:
            print(f"[Error] JSON导出失败: {e}")
            return False


def main():
    """主函数"""
    import sys
    
    exporter = DataExporter()
    
    # 导出为Excel（或CSV降级）
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    excel_path = f'/root/.openclaw/workspace/portfolio-tracker/exports/portfolio_export_{timestamp}.xlsx'
    
    # 确保目录存在
    import os
    os.makedirs(os.path.dirname(excel_path), exist_ok=True)
    
    if exporter.export_to_excel(excel_path):
        print(f"\n📁 导出文件: {excel_path}")
        return 0
    else:
        return 1


if __name__ == '__main__':
    exit(main())
